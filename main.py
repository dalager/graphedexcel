from openpyxl import load_workbook


def extract_formulas(file_path):
    # Load the workbook with formulas (data_only=False ensures we're loading formulas, not just the results)
    wb = load_workbook(file_path, data_only=False)

    # Iterate over all worksheets in the workbook
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n-- Extracting formulas from sheet: {sheet} --\n")

        # Iterate over all cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    # Print the sheet, cell location, and formula
                    print(
                        f"Sheet: {sheet}, Cell: {cell.coordinate}, Formula: {cell.value}"
                    )


if __name__ == "__main__":
    # Replace 'your_spreadsheet.xlsx' with the path to your Excel file
    path_to_excel = "simplified_1.xlsx"
    extract_formulas(path_to_excel)
