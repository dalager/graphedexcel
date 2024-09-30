"""
This script extracts formulas from an Excel file and builds a dependency graph.
"""

from typing import List, Dict
from openpyxl import load_workbook
import networkx as nx
import re
import sys
from .graph_visualizer import visualize_dependency_graph
from .graph_summarizer import print_summary
from .excel_parser import extract_references

# Dictionary that stores the unique functions used in the formulas
# The key will be the function name and the value will be the number of times it was used
functions_dict: Dict[str, int] = {}


def log(msg: str) -> None:
    """
    Log a message to the console if verbosity is enabled using the --verbose flag.
    """
    if "--verbose" in sys.argv:
        print(msg)


def sanitize_sheetname(sheetname: str) -> str:
    """
    Remove any special characters from the sheet name.
    """
    return sheetname.replace("'", "")


def sanitize_range(rangestring: str) -> str:
    """
    Remove any special characters from the range.
    """
    if "!" in rangestring:
        sheet, range_ = rangestring.split("!")
        sheet = sheet.replace("'", "")
        return f"{sheet}!{range_}"
    return rangestring


def stat_functions(cellvalue: str) -> None:
    """
    Extract the functions used in the formula and store them in a dictionary.
    This will be used to print the most used functions in the formulas.
    """
    cellfuncs = re.findall(r"[A-Z]+\(", cellvalue)
    log(f"  Functions used: {functions_dict}")
    for function in cellfuncs:
        function = function[:-1]  # Remove the "(" from the function name
        functions_dict[function] = functions_dict.get(function, 0) + 1


def add_node(graph: nx.DiGraph, node: str, sheet: str) -> None:
    """
    Add a node to the graph with the specified sheet name.
    """
    log(f"Adding node: {node} in sheet: {sheet}")
    sheet = sanitize_sheetname(sheet)
    graph.add_node(node, sheet=sheet)


def extract_formulas_and_build_dependencies(file_path: str) -> nx.DiGraph:
    """
    Extract formulas from an Excel file and build a dependency graph.
    """
    try:
        wb = load_workbook(file_path, data_only=False)
    except Exception as e:
        log(f"Error loading workbook: {e}")
        sys.exit(1)

    graph = nx.DiGraph()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log(f"========== Analyzing sheet: {sheet_name} ==========")
        sanitized_sheet_name = sanitize_sheetname(sheet_name)
        process_sheet(ws, sanitized_sheet_name, graph)

    return graph


def process_sheet(ws, sheet_name: str, graph: nx.DiGraph) -> None:
    """
    Process a sheet and add references to the graph.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                process_formula_cell(cell, sheet_name, graph)


def process_formula_cell(cell, sheet_name: str, graph: nx.DiGraph) -> None:
    """
    Process a cell containing a formula.
    """
    stat_functions(cell.value)
    cell_reference = f"{sheet_name}!{cell.coordinate}"
    log(f"Formula in {cell_reference}: {cell.value}")
    add_node(graph, cell_reference, sheet_name)

    direct_references, range_references, range_dependencies = extract_references(
        cell.value
    )
    add_references_to_graph(direct_references, cell_reference, sheet_name, graph)
    add_ranges_to_graph(range_references, cell_reference, sheet_name, graph)
    add_range_dependencies_to_graph(range_dependencies, sheet_name, graph)


def add_references_to_graph(
    references: List[str], current_cell: str, sheet_name: str, graph: nx.DiGraph
) -> None:
    """
    Add direct cell references to the graph.
    """
    for cell_reference in references:
        cell_reference = format_reference(cell_reference, sheet_name)
        log(f"  Cell: {cell_reference}")
        add_node(graph, cell_reference, sheet_name)
        graph.add_edge(current_cell, cell_reference)


def add_ranges_to_graph(
    ranges: List[str], current_cell: str, sheet_name: str, graph: nx.DiGraph
) -> None:
    """
    Add range references to the graph.
    """
    for range_reference in ranges:
        range_sheet_name = get_range_sheet_name(range_reference, sheet_name)
        range_reference = format_reference(range_reference, sheet_name)
        log(f"  Range: {range_reference}")
        add_node(graph, range_reference, range_sheet_name)
        graph.add_edge(current_cell, range_reference)


def add_range_dependencies_to_graph(
    range_dependencies: Dict[str, str], sheet_name: str, graph: nx.DiGraph
) -> None:
    """
    Add dependencies between ranges and cells.
    """
    for cell_reference, range_reference in range_dependencies.items():
        range_reference = format_reference(range_reference, sheet_name)
        cell_reference = format_reference(cell_reference, sheet_name)
        range_sheet_name = range_reference.split("!")[0]
        cell_sheet_name = cell_reference.split("!")[0]

        add_node(graph, cell_reference, cell_sheet_name)
        add_node(graph, range_reference, range_sheet_name)
        graph.add_edge(range_reference, cell_reference)


def format_reference(reference: str, sheet_name: str) -> str:
    """
    Format a cell or range reference to include the sheet name if not already present.
    """
    return (
        f"{sheet_name}!{reference}"
        if "!" not in reference
        else reference.replace("'", "")
    )


def get_range_sheet_name(range_reference: str, sheet_name: str) -> str:
    """
    Get the sheet name for a range reference.
    """
    return sheet_name if "!" not in range_reference else range_reference.split("!")[0]




if __name__ == "__main__":
    path_to_excel = "Book1.xlsx"

    if len(sys.argv) > 1:
        path_to_excel = sys.argv[1]

    # Extract formulas and build the dependency graph
    dependency_graph = extract_formulas_and_build_dependencies(path_to_excel)

    print_summary(dependency_graph, functions_dict)

    # if --no-visualize is not passed as argument
    if "--no-visualize" not in sys.argv:
        print(
            "\033[1;30;40m\nVisualizing the graph of dependencies.\nThis might take a while...\033[0;37;40m\n"  # noqa
        )

        visualize_dependency_graph(dependency_graph, path_to_excel)