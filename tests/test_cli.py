import os
import tempfile
from openpyxl import Workbook
import pytest
import sys

# from argparse import Namespace
from graphedexcel.cli import parse_arguments, main
from unittest.mock import patch, MagicMock


@pytest.fixture
def create_excel_file(tmp_path):
    def _create_excel_file(data):
        file_path = tmp_path / "test.xlsx"
        wb = Workbook()
        for sheet_name, sheet_data in data.items():
            ws = wb.create_sheet(title=sheet_name)
            for row in sheet_data:
                ws.append(row)
        wb.save(file_path)
        return file_path

    return _create_excel_file


# test cli.main
def test_main():
    # assert that main with no arguments raises SystemExit error
    test_args = ["graphedexcel"]
    with patch("sys.argv", test_args):
        with pytest.raises(SystemExit):
            main()


def test_main_with_nonexistent_file(capsys):
    test_args = ["graphedexcel", "nonexistent_file.xlsx"]
    with patch("sys.argv", test_args):
        with pytest.raises(SystemExit) as exc_info:
            main()
        assert exc_info.value.code != 0
    captured = capsys.readouterr()
    assert "File not found:" in captured.err


def test_main_with_test_xlsx_file(capsys):
    """Test main with a test excel file to exercise the cli module"""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        test_file_path = tmp_file.name
        try:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Test Data"

            wb.save(test_file_path)
            wb.close()
            test_args = ["graphedexcel", test_file_path]
            with patch("sys.argv", test_args):
                main()
        finally:
            print("here")
            wb.close()
    captured = capsys.readouterr()
    assert "Dependency graph image saved" in captured.out


def test_parse_arguments_required(monkeypatch):
    """
    Test that the required positional argument is parsed correctly.
    """
    test_args = ["graphedexcel", "test.xlsx"]
    with patch("sys.argv", test_args):
        args = parse_arguments()
        assert args.path_to_excel == "test.xlsx"


def test_parse_arguments_optional_flags():
    """
    Test that optional flags are parsed correctly.
    """
    test_args = [
        "graphedexcel",
        "test.xlsx",
        "--as-directed-graph",
        "--no-visualize",
        "--hide-legends",
    ]
    with patch("sys.argv", test_args):
        args = parse_arguments()
        assert args.path_to_excel == "test.xlsx"
        assert args.as_directed_graph is True
        assert args.no_visualize is True
        assert args.hide_legends is True


def test_parse_arguments_optional_arguments():
    """
    Test that optional arguments are parsed correctly.
    """
    test_args = [
        "graphedexcel",
        "test.xlsx",
        "--layout",
        "circular",
        "--config",
        "config.json",
        "--output-path",
        "output.png",
        "--open-image",
    ]
    with patch("sys.argv", test_args):
        args = parse_arguments()
        assert args.path_to_excel == "test.xlsx"
        assert args.layout == "circular"
        assert args.config == "config.json"
        assert args.output_path == "output.png"
        assert args.open_image is True


def test_parse_arguments_default_values():
    """
    Test that default values are set correctly.
    """
    test_args = ["graphedexcel", "test.xlsx"]
    with patch("sys.argv", test_args):
        args = parse_arguments()
        assert args.layout == "spring"
        assert args.config is None
        assert args.output_path is None
        assert args.as_directed_graph is False
        assert args.no_visualize is False
        assert args.open_image is False
        assert args.hide_legends is None


def test_parse_arguments_invalid():
    """
    Test that invalid arguments raise a SystemExit.
    """
    test_args = ["graphedexcel"]
    with patch("sys.argv", test_args):
        with pytest.raises(SystemExit):
            parse_arguments()
