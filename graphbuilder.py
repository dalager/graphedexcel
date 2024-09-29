"""
This script extracts formulas from an Excel file and builds a dependency graph.
"""

from openpyxl import load_workbook
from collections import Counter
import networkx as nx
import re
import sys
from graph_visualizer import visualize_dependency_graph
from excel_parser import extract_references

# dictionary that stores the uniqe functions used in the formulas
# the key will be the funciton name and the value will be the number of times it was used
functions_dict = {}


def log(msg):
    """
    Log a message to the console if verbosity is enabled using the --verbose flag.
    """
    # if verbosity is enabled

    if "--verbose" in sys.argv:
        print(msg)


def sanitize_sheetname(sheetname):
    """
    Remove any special characters from the sheet name.
    """
    return sheetname.replace("'", "")


def sanitize_range(rangestring):
    """
    Remove any special characters from the range.
    """
    if "!" in rangestring:
        sheet = rangestring.split("!")[0].replace("'", "")
        range = rangestring.split("!")[1]

    return f"{sheet}!{range}"


def stat_functions(cellvalue):
    """
    Extract the functions used in the formula and store them in a dictionary.
    This will be used to print the most used functions in the formulas.
    """

    # functions used in the formula
    cellfuncs = re.findall(r"[A-Z]+\(", cellvalue)
    log(f"  Functions used: {functions_dict}")
    for function in cellfuncs:
        # remove the "(" from the function name
        function = function[:-1]
        if function in functions_dict:
            functions_dict[function] += 1
        else:
            functions_dict[function] = 1


def add_node(graph, node, sheet):
    """
    Add a node to the graph with the specified sheet name.
    """
    log(f"Adding node: {node} in sheet: {sheet}")
    sheet = sanitize_sheetname(sheet)
    graph.add_node(node, sheet=sheet)


def extract_formulas_and_build_dependencies(file_path):
    """
    Extract formulas from an Excel file and build a dependency graph.
    """
    wb = load_workbook(file_path, data_only=False)
    graph = nx.DiGraph()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log(f"========== Analyzing sheet: {sheet_name} ==========")
        sanitized_sheet_name = sanitize_sheetname(sheet_name)
        process_sheet(ws, sanitized_sheet_name, graph)

    return graph


def process_sheet(ws, sheet_name, graph):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                process_formula_cell(cell, sheet_name, graph)


def process_formula_cell(cell, sheet_name, graph):
    stat_functions(cell.value)
    cell_reference = f"{sheet_name}!{cell.coordinate}"
    log(f"Formula in {cell_reference}: {cell.value}")
    add_node(graph, cell_reference, sheet_name)

    direct_references, range_references, range_dependencies = extract_references(
        cell.value
    )
    add_references_to_graph(direct_references, cell_reference, sheet_name, graph)
    add_ranges_to_graph(range_references, cell_reference, sheet_name, graph)
    add_range_dependencies_to_graph(range_dependencies, sheet_name, graph)


def add_references_to_graph(references, current_cell, sheet_name, graph):
    for cell_reference in references:
        cell_reference = (
            f"{sheet_name}!{cell_reference}"
            if "!" not in cell_reference
            else cell_reference.replace("'", "")
        )
        log(f"  Cell: {cell_reference}")
        add_node(graph, cell_reference, sheet_name)
        graph.add_edge(current_cell, cell_reference)


def add_ranges_to_graph(ranges, current_cell, sheet_name, graph):
    for range_reference in ranges:
        range_sheet_name = (
            sheet_name if "!" not in range_reference else range_reference.split("!")[0]
        )
        range_reference = (
            f"{sheet_name}!{range_reference}"
            if "!" not in range_reference
            else sanitize_range(range_reference)
        )
        log(f"  Range: {range_reference}")
        add_node(graph, range_reference, range_sheet_name)
        graph.add_edge(current_cell, range_reference)


def add_range_dependencies_to_graph(range_dependencies, sheet_name, graph):
    for cell_reference, range_reference in range_dependencies.items():
        range_reference = (
            f"{sheet_name}!{range_reference}"
            if "!" not in range_reference
            else sanitize_range(range_reference)
        )
        cell_reference = (
            f"{sheet_name}!{cell_reference}"
            if "!" not in cell_reference
            else cell_reference
        )
        range_sheet_name = range_reference.split("!")[0]
        cell_sheet_name = cell_reference.split("!")[0]

        add_node(graph, cell_reference, cell_sheet_name)
        add_node(graph, range_reference, range_sheet_name)
        graph.add_edge(range_reference, cell_reference)


def print_summary(graph, functionsdict):
    """
    Summarize a networkx DiGraph representing a dependency graph. And print the most used functions in the formulas
    """

    strpadsize = 28
    numpadsize = 5
    # 1. Print basic information about the graph

    print("=== Dependency Graph Summary ===")
    print(
        "Cell/Node count".ljust(strpadsize, " ")
        + str(graph.number_of_nodes()).rjust(numpadsize, " ")
    )
    print(
        "Dependency count".ljust(strpadsize, " ")
        + str(graph.number_of_edges()).rjust(numpadsize, " ")
    )
    print()

    # 2. Print the nodes with the highest degree
    degree_view = graph.degree()

    degree_counts = Counter(dict(degree_view))
    max_degree_node = degree_counts.most_common(10)
    print("=== Nodes with the highest degree ===")
    for node, degree in max_degree_node:
        print(f"{node.ljust(strpadsize)}{str(degree).rjust(numpadsize, ' ')} ")

    # 3. Print the most used functions
    print("\n=== Formula functions by count ===")
    sorted_functions = dict(
        sorted(functionsdict.items(), key=lambda item: item[1], reverse=True)
    )

    for function, count in sorted_functions.items():
        print(f"{function.ljust(strpadsize, ' ')}{str(count).rjust(numpadsize, ' ')}")


if __name__ == "__main__":
    path_to_excel = "Book1.xlsx"

    if len(sys.argv) > 1:
        path_to_excel = sys.argv[1]

    # Extract formulas and build the dependency graph
    dependency_graph = extract_formulas_and_build_dependencies(path_to_excel)

    print_summary(dependency_graph, functions_dict)

    # if --no-visualize is not passed as argument
    if "--no-visualize" not in sys.argv:
        print(
            "\033[1;30;40m\nVisualizing the graph of dependencies.\nThis might take a while...\033[0;37;40m\n"  # noqa
        )

        visualize_dependency_graph(dependency_graph, path_to_excel)
