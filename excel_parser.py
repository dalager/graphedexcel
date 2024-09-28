from openpyxl.utils import get_column_letter, range_boundaries
import re


# Regex to detect cell references like A1, B2, or ranges like A1:B2
CELL_REF_REGEX = r"('?[A-Za-z0-9_\-\[\] ]+'?![A-Z]{1,3}[0-9]+(:[A-Z]{1,3}[0-9]+)?)|([A-Z]{1,3}[0-9]+(:[A-Z]{1,3}[0-9]+)?)"  # noqa


def expand_range(range_ref):
    """
    Expand a range reference (e.g., 'A1:A3') into a list of individual cell references.
    """

    # if there is a sheet name in the range reference, put it away for now
    if "!" in range_ref:
        sheet_name, range_ref = range_ref.split("!")
    else:
        sheet_name = None

    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    expanded_cells = []

    # Loop over rows and columns in the range
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            # if sheetname is set
            if sheet_name:
                expanded_cells.append(f"{sheet_name}!{get_column_letter(col)}{row}")
            else:
                expanded_cells.append(f"{get_column_letter(col)}{row}")

    return expanded_cells


def extract_references(formula):
    """
    Extract all referenced cells and ranges from a formula using regular expressions.
    This returns a list of both individual cells and range references.
    """
    formula = formula.replace("$", "")
    matches = re.findall(CELL_REF_REGEX, formula)
    references = [match[0] if match[0] else match[2] for match in matches]

    # trim the extracted references
    references = [ref.strip() for ref in references]

    expanded_references = []
    dependencies = {}
    direct_references = []
    range_references = []

    for ref in references:
        if ":" in ref:  # it's a range like A1:A3
            expanded_cells = expand_range(ref)
            expanded_references.extend(expanded_cells)
            range_references.append(ref)
            # Store the range-to-cells relationship
            for cell in expanded_cells:
                dependencies[cell] = ref
        else:  # single cell
            direct_references.append(ref)

    return direct_references, range_references, dependencies
